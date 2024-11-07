import os
import zipfile

from openpyxl import load_workbook
from pypdf import PdfReader

current_file = os.path.abspath(__file__)
current_dir = os.path.dirname(current_file)
path_to_directory_files = os.path.join(current_dir, "files")
path_to_directory_archive = os.path.join(current_dir, "arhives")


def test_read_pdf(add_files_to_zip):
    with zipfile.ZipFile(os.path.join(path_to_directory_archive, "test.zip"), "r") as test_zip:
        with test_zip.open("test.pdf", "r") as pdf:
            reader = PdfReader(pdf).pages
            expected_pdf_text = ("Никакой полезной информации он не несёт.")
            assert expected_pdf_text in reader[0].extract_text()


def test_read_xlsx(add_files_to_zip):
    with zipfile.ZipFile(os.path.join(path_to_directory_archive, "test.zip"), "r") as test_zip:
        with test_zip.open("test.xlsx", "r") as xlsx:
            workbook = load_workbook(xlsx).active
            assert workbook.cell(row=20, column=3).value == 1


def test_read_csv(add_files_to_zip):
    with zipfile.ZipFile(os.path.join(path_to_directory_archive, "test.zip"), "r") as test_zip:
        content = test_zip.read("test.csv")
        assert "Ms. Odessa Hartmann" in str(content)
